import datetime
import cStringIO

import os
import re
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl.writer.excel as wrtex
from os import listdir
from os.path import isfile, join


def get_all_values_from_ws(ws, d):
    """return an array of arrays f.e column in a ws"""

    if d == 'r':
        return [get_values(r) for r in ws.rows if r[0].value is not None]

    if d == 'c':
        return [get_values(c) for c in ws.columns if c[0].value is not None]


def pull_wb(location, src, strip):
    """return an excel file from either local or source"""
    if strip:
        w = wb_strip(location, src)
    else:
        if src == 'db':
            w = load_workbook(pull_from_db(location))
        else:
            w = load_workbook(location)

    return w

def wb_strip(location, src):
    if src == 'db':
        w = load_workbook(pull_from_db(location), read_only = True, data_only = True)
    else:
        w = load_workbook(location, read_only = True, data_only = True)

    new_wb = Workbook()
    if len(new_wb.worksheets) > 0:
        new_wb.remove_sheet(new_wb.worksheets[0])

    for v in w.worksheets:
        cur_w = new_wb.create_sheet(1, v.title)
        for r in v.rows:
            for v in r:
                if v.value is not None:
                    cur_w[v.coordinate] = v.value

    print "Pulled: " + location
    print "With tabs: " + str(new_wb.get_sheet_names())

    return new_wb

def pull_from_db(path):
    """pull a file from dropbox"""
    to_ret = cStringIO.StringIO()

    with client.get_file(path) as f:
        to_ret.write(f.read())
    f.close()

    return to_ret

def get_file_list(path, src):
    """return file list from local or db"""
    if src == 'db':
        meta = client.metadata(path, list=True)
        return [str(f['path']) for f in meta['contents'] if re.search('xls|xlsx$',str(f))]

    elif src == 'local':
        return [str(path +'/' + f) for f in listdir(path) if isfile(join(path,f)) and re.search('xls|xlsx$',str(f))]

def fuzzy_match_col(col, vals):
    pass

def xstr(conv, **kargs):
    """return a more battle teststed encoded string"""
    if (conv == 'None' or conv is None or conv == '') and kargs.has_key('setnull') and kargs['setnull']:
        return None

    try:
        return str(conv.encode('utf8'))
    except:
        return str(conv)


def get_values(r, **kargs):
    """returns values of a row or a column - note dates are specifically formatted"""
    #TODO: parameterize date format?
    ret = []
    for v in r:
        if xstr(v.value) == 'None':
            if kargs.has_key('setnull') and kargs['setnull']:
                ret.append(None)
            else:
                ret.append('')
        elif isinstance(v.value, datetime.datetime):
            ret.append(v.value.strftime('%d/%m/%Y'))
        else:
            ret.append(xstr(v.value))

    return ret

def send_wb(path, wb, src):
    print 'Sending... ' + path
    if src == 'db':
        client.put_file(path, wrtex.save_virtual_workbook(wb))

    elif src == 'local':
        if not os.path.exists(path.rsplit('/', 1)[0]):
            os.makedirs(path.rsplit('/', 1)[0])
        wb.save(path)

    print 'Sent!'
