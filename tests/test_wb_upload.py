import unittest
from exql import wb_upload
from exql.Sheet import Sheet
from sqlalchemy import Column, Integer, String, Float
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import relationship
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import column_index_from_string
import os

Base = declarative_base()
engine = create_engine(os.environ['dbk'])
Base.metadata.bind = engine
DBSession = sessionmaker(bind=engine)
session = DBSession()
Base.metadata.create_all(engine)


class TestUpload(unittest.TestCase):

    def get_ws(self):
        w = Workbook()
        w.active.title = 'new'
        w.active.append(("Column 1", "Column 2"))
        w.active.append(("v1", "v2"))
        w.active.append(("v11", "v22"))

        return w

    def setUp(self):
        self.tw = self.get_ws()

    def test_append(self):
        wb = Workbook()
        ret = xls_to_sql.get_sheets([wb], ('sheet1', 'sheet3'))
        self.assertEqual([v.title for v in xls_to_sql.get_sheets([wb], ('sheet1', 'sheet3'))] \
                         , ['sheet1','sheet3'])

    def test_create_with_sheet(self):
        sht = Sheet(ws = self.tw.get_sheet_by_name('new'), name = 'upload_test', col_nms = ['col1','col2'])
        e = create_engine(os.environ['dbk'])
        #e = create_engine('sqlite://')
        wb_upload.create_with_sheet(sht, e)


    def test_get_sht_nm_no_param(self):
        self.assertEqual(upload_a_wb.get_sht_nm(self.tw, None), 'new')



class test(Base):
    __tablename__ = 'test'
    targeting = Column(String(250), primary_key=True)
    quantity = Column(Integer)
    total_hh = Column(Integer)
    float = Column(Float)

if __name__ == '__main__':
    unittest.main()
